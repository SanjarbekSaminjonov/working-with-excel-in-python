import datetime
import pytz

from openpyxl import load_workbook
from openpyxl.workbook import Workbook


def get_str_datetime() -> str:
    tz = pytz.timezone('Asia/Tashkent')
    return str(datetime.datetime.now(tz=tz))


def str_to_datetime(str_datetime: str) -> datetime.datetime:
    return datetime.datetime.strptime(str_datetime, '%Y-%m-%d %H:%M:%S.%f%z')


class Excel:
    HOME_PAGE = 'Bosh sahifa'
    UPDATED = 'Yangilangan'
    UPDATE_INFO_CELL = 'A1'
    UPDATE_INFO_CELL_DATETIME = 'A2'

    def __init__(self, file_name: str):
        self.file_name = f'{file_name}.xlsx'
        try:
            self.wb = load_workbook(self.file_name)
            self.ws = self.wb[self.HOME_PAGE]
        except FileNotFoundError:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = self.HOME_PAGE

    def get_last_update(self) -> datetime.datetime:
        home_page = self.wb[self.HOME_PAGE]
        last_update = home_page[self.UPDATE_INFO_CELL_DATETIME].value
        return str_to_datetime(last_update) if last_update else None

    def get_file_name(self) -> str:
        return self.file_name

    def set_sheet(self, sheet_name: str, header: list):
        try:
            self.ws = self.wb[sheet_name]
        except KeyError:
            self.ws = self.wb.create_sheet(sheet_name)
            self.ws.append(header)

    def add_row(self, row: list):
        self.ws.append(row)

    def save(self):
        home_page = self.wb[self.HOME_PAGE]
        home_page[self.UPDATE_INFO_CELL].value = self.UPDATED
        home_page[self.UPDATE_INFO_CELL_DATETIME] = get_str_datetime()
        self.wb.save(self.file_name)


ex = Excel('myfile')

# Last update of the file
print(ex.get_last_update())

# Wordl sheet
ex.set_sheet('World', ['Message', 'DateTime'])
ex.add_row(['Hello World!', get_str_datetime()])
ex.add_row(['Good Bye World!', get_str_datetime()])

# Uzbekistan sheet
ex.set_sheet('Uzbekiston', ['Message', 'From', 'DateTime'])
ex.add_row(['Hello Uzbekistan!', 'Sanjarbek', get_str_datetime()])
ex.add_row(['Good Bye Uzbekistan!', 'Shukurullo', get_str_datetime()])

# Save the file
ex.save()
