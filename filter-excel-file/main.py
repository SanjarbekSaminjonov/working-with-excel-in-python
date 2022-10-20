from openpyxl import load_workbook
from openpyxl.workbook import Workbook

from collections import defaultdict
from datetime import datetime


class Excel:

    def __init__(self, file_name: str):
        self.file_name = f'{file_name}.xlsx'
        try:
            self.wb = load_workbook(self.file_name)
        except FileNotFoundError:
            self.wb = Workbook()
        self.ws = self.wb.active

    def get_row_len(self):
        return self.ws.max_row

    def get_row(self, row_number: int):
        return self.ws[row_number]

    def set_sheet(self, sheet_name: str):
        try:
            self.ws = self.wb[sheet_name]
        except KeyError:
            self.ws = self.wb.create_sheet(sheet_name)

    def add_row(self, row_elements: list):
        self.ws.append(row_elements)

    def save(self):
        self.wb.save(self.file_name)


now = datetime.now()
years = defaultdict(lambda: [])

ex = Excel('Book1')
new_ex = Excel('New')

for i in range(3, ex.get_row_len() + 1):
    value = ex.ws[f'E{i}'].value
    if type(value) == type(now):
        years[value.year].append(ex.ws[i])
    elif value is not None:
        year = value[-4:]
        try:
            years[int(year)].append(ex.ws[i])
        except ValueError or TypeError:
            years[0].append(ex.ws[i])
    else:
        years[0].append(ex.ws[i])

for key in sorted(years.keys(), reverse=True):
    list_ = years[key]
    new_ex.set_sheet(str(key))
    new_ex.add_row(
        ['№', 'Хонадон сони', 'Оила раками', 'Ф.И.Ш.', 'Ой Кун Йил', 'Хужжат раками', 'Пнифл', 'Манзил', 'Изоҳ'])
    for row_id in range(1, len(list_) + 1):
        row_items = []
        row = list(list_[row_id - 1])
        row[0].value = row_id
        new_ex.add_row([cell.value for cell in row])

new_ex.save()
